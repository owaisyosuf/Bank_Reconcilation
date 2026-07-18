"""Color-coded Excel export (M4).

Builds the reconciliation workbook described in CLAUDE.md's UI Requirements
and REQUIREMENTS.md F4.4/F4.5:

- Sheet 1 "Summary": counts + total amounts per tier, matched %, generated-at
  timestamp.
- One sheet per tier ("Exact", "Tolerance", "Review", "Bank Only",
  "Ledger Only"): bank + ledger date/description/debit/credit, amount_diff,
  date_diff_days, description_score, reason, and both source_row values for
  traceability.

Pure formatting/serialization layer: never re-implements matching or parsing
logic (ui-agent.md rule 8). All PII masking goes through the shared
`app.ui.pii.mask_pii()` helper (CLAUDE.md Hard Rule #5). Decimal->float
conversion happens ONLY here, at the final xlsx-cell-value boundary, because
openpyxl cells don't support Decimal directly.
"""

from __future__ import annotations

import datetime
import io
from decimal import Decimal
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.matching.engine import ReconciliationResult
from app.matching.scoring import MatchRecord
from app.parsers.base import StandardTransaction
from app.ui.pii import mask_pii

# Exact same hex values as CLAUDE.md / ui-agent.md rule 3 and
# app/ui/main.py's TIER_COLORS, so the workbook and the on-screen tables
# never disagree.
TIER_COLORS = {
    "EXACT": "C6EFCE",
    "TOLERANCE": "FFEB9C",
    "REVIEW": "FFD8B1",
    "UNMATCHED": "FFC7CE",
}

_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_HEADER_FONT = Font(bold=True)

_MAX_COL_WIDTH = 50
_MIN_COL_WIDTH = 8

DETAIL_COLUMNS = [
    "Bank Date",
    "Bank Description",
    "Bank Debit",
    "Bank Credit",
    "Ledger Date",
    "Ledger Description",
    "Ledger Debit",
    "Ledger Credit",
    "Amount Diff",
    "Date Diff (days)",
    "Description Score",
    "Reason",
    "Bank Row #",
    "Ledger Row #",
]

# (sheet title, attribute on ReconciliationResult, fill color key)
_TIER_SHEETS = [
    ("Exact", "exact", "EXACT"),
    ("Tolerance", "tolerance", "TOLERANCE"),
    ("Review", "review", "REVIEW"),
    ("Bank Only", "bank_only", "UNMATCHED"),
    ("Ledger Only", "ledger_only", "UNMATCHED"),
]


def _to_float(value: Optional[Decimal]) -> Optional[float]:
    """Decimal -> float, only for the purpose of writing an xlsx numeric
    cell. Never used anywhere in matching/parsing (see module docstring)."""
    if value is None:
        return None
    return float(value)


def _txn_amount(txn: Optional[StandardTransaction]) -> Decimal:
    if txn is None:
        return Decimal("0")
    return txn.credit - txn.debit


def _autofit_columns(ws: Worksheet, num_columns: int) -> None:
    """openpyxl has no true autofit; approximate it from the max rendered
    length of each column's contents, capped to keep sheets usable."""
    for col_idx in range(1, num_columns + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            if length > max_len:
                max_len = length
        width = min(max(max_len + 2, _MIN_COL_WIDTH), _MAX_COL_WIDTH)
        ws.column_dimensions[letter].width = width


def _write_header(ws: Worksheet, columns: list[str]) -> None:
    for col_idx, title in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _record_to_row(record: MatchRecord) -> list:
    bank, ledger = record.bank_txn, record.ledger_txn
    return [
        bank.date.isoformat() if bank else "",
        mask_pii(bank.description) if bank else "",
        _to_float(bank.debit) if bank else None,
        _to_float(bank.credit) if bank else None,
        ledger.date.isoformat() if ledger else "",
        mask_pii(ledger.description) if ledger else "",
        _to_float(ledger.debit) if ledger else None,
        _to_float(ledger.credit) if ledger else None,
        _to_float(record.amount_diff),
        record.date_diff_days,
        record.description_score,
        record.reason,
        bank.source_row if bank else None,
        ledger.source_row if ledger else None,
    ]


def _write_tier_sheet(wb: Workbook, title: str, records: list[MatchRecord], color_hex: str) -> None:
    ws = wb.create_sheet(title=title)
    _write_header(ws, DETAIL_COLUMNS)

    fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
    for row_idx, record in enumerate(records, start=2):
        row_values = _record_to_row(record)
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill

    _autofit_columns(ws, len(DETAIL_COLUMNS))


def _tier_total_amount(records: list[MatchRecord]) -> Decimal:
    total = Decimal("0")
    for record in records:
        # Prefer the bank-side amount (present for every tier except
        # ledger-only, where only the ledger side exists).
        source = record.bank_txn if record.bank_txn is not None else record.ledger_txn
        total += abs(_txn_amount(source))
    return total


def _write_summary_sheet(wb: Workbook, result: ReconciliationResult) -> None:
    ws = wb.create_sheet(title="Summary", index=0)

    columns = ["Metric", "Count", "Total Amount (Rs)"]
    _write_header(ws, columns)

    tier_data = [
        ("Exact", result.exact, "EXACT"),
        ("Tolerance", result.tolerance, "TOLERANCE"),
        ("Review", result.review, "REVIEW"),
        ("Bank Only (unmatched)", result.bank_only, "UNMATCHED"),
        ("Ledger Only (unmatched)", result.ledger_only, "UNMATCHED"),
    ]

    row_idx = 2
    for label, records, color_key in tier_data:
        fill = PatternFill(
            start_color=TIER_COLORS[color_key], end_color=TIER_COLORS[color_key], fill_type="solid"
        )
        ws.cell(row=row_idx, column=1, value=label).fill = fill
        ws.cell(row=row_idx, column=2, value=len(records)).fill = fill
        ws.cell(row=row_idx, column=3, value=_to_float(_tier_total_amount(records))).fill = fill
        row_idx += 1

    row_idx += 1  # blank spacer row

    total_bank_side = (
        len(result.exact) + len(result.tolerance) + len(result.review) + len(result.bank_only)
    )
    matched_pct = (result.matched_count / total_bank_side * 100) if total_bank_side else 0.0
    total_matched_amount = (
        _tier_total_amount(result.exact) + _tier_total_amount(result.tolerance) + _tier_total_amount(result.review)
    )

    ws.cell(row=row_idx, column=1, value="Matched %")
    ws.cell(row=row_idx, column=2, value=round(matched_pct, 2))
    row_idx += 1

    ws.cell(row=row_idx, column=1, value="Total Matched Amount (Rs)")
    ws.cell(row=row_idx, column=2, value=_to_float(total_matched_amount))
    row_idx += 1

    ws.cell(row=row_idx, column=1, value="Total Matched Count")
    ws.cell(row=row_idx, column=2, value=result.matched_count)
    row_idx += 1

    ws.cell(row=row_idx, column=1, value="Total Unmatched Count")
    ws.cell(row=row_idx, column=2, value=result.unmatched_count)
    row_idx += 1

    ws.cell(row=row_idx, column=1, value="Generated At")
    ws.cell(
        row=row_idx,
        column=2,
        value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    )

    _autofit_columns(ws, len(columns))


def build_excel_report(result: ReconciliationResult) -> bytes:
    """Build the full color-coded reconciliation workbook and return it as
    bytes, ready for `st.download_button(data=...)`.

    Sheet order: Summary, Exact, Tolerance, Review, Bank Only, Ledger Only.
    """
    wb = Workbook()
    # Drop the default blank sheet openpyxl creates; we build our own set.
    default_sheet = wb.active
    wb.remove(default_sheet)

    _write_summary_sheet(wb, result)

    for title, attr, color_key in _TIER_SHEETS:
        records = getattr(result, attr)
        _write_tier_sheet(wb, title, records, TIER_COLORS[color_key])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
