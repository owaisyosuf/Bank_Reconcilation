"""Streamlit entrypoint for the bank reconciliation app (M3).

Single-page app: sidebar for uploads/settings, main area for results.
This module owns UI wiring only — parsing lives in `app/parsers/`, matching
lives in `app/matching/`. Nothing here re-implements that logic (CLAUDE.md
Hard Rules; ui-agent.md rule 8).

Run with:
    streamlit run app/ui/main.py
"""

from __future__ import annotations

import csv
import datetime
import hashlib
import io
from decimal import Decimal, InvalidOperation

import pandas as pd
import streamlit as st

try:
    import openpyxl
except ImportError:  # pragma: no cover - openpyxl is a hard dependency
    openpyxl = None

from app.matching.engine import ReconciliationResult, reconcile
from app.matching.scoring import MatchConfig, MatchRecord
from app.parsers import BANK_ADAPTERS, get_parser
from app.parsers.base import ScannedPdfError, StandardTransaction
from app.parsers.ledger_parser import LedgerParser
from app.ui.pii import mask_pii

SCANNED_PDF_MESSAGE = (
    "Ye PDF scan ki hui image hai. Bank portal se CSV ya Excel download kar "
    "ke upload karein. (This PDF appears to be a scanned image with no "
    "extractable text — please download the statement as CSV or Excel from "
    "your bank's portal instead.)"
)

TIER_COLORS = {
    "EXACT": "#C6EFCE",
    "TOLERANCE": "#FFEB9C",
    "REVIEW": "#FFD8B1",
    "UNMATCHED": "#FFC7CE",
}

NONE_OPTION = "(none)"
MODE_SEPARATE = "Separate Debit & Credit columns"
MODE_AMOUNT_DRCR = "Single Amount column + DR/CR column"

TABLE_COLUMNS = [
    "Tier",
    "Bank Date",
    "Bank Description",
    "Bank Debit",
    "Bank Credit",
    "Ledger Date",
    "Ledger Description",
    "Ledger Debit",
    "Ledger Credit",
    "Amount Diff",
    "Date Diff (days)",
    "Desc Score",
    "Reason",
    "Bank Row #",
    "Ledger Row #",
]


# ---------------------------------------------------------------------------
# Raw file reading (UI-layer concern only: populating the column-mapping
# dropdowns and previews). All *interpretation* of the data — date/amount
# parsing, debit/credit normalization — is still done by app/parsers/.
# ---------------------------------------------------------------------------


def _read_raw_rows(file_bytes: bytes, filename: str) -> list[list]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        text = file_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        return [list(row) for row in reader]

    if lower.endswith((".xlsx", ".xlsm", ".xls")):
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        sheet = wb[wb.sheetnames[0]]
        return [list(row) for row in sheet.iter_rows(values_only=True)]

    raise ValueError(f"Unsupported ledger file type: {filename}")


def _header_labels(header_row: list) -> list[str]:
    labels = []
    for i, cell in enumerate(header_row):
        text = "" if cell is None else str(cell).strip()
        labels.append(text if text else f"col_{i}")
    return labels


def _build_amount_drcr_csv(
    rows: list[list],
    idx: dict[str, int],
    description_key: str | None,
    balance_key: str | None,
) -> tuple[bytes, dict[str, str]]:
    """Reshape a "single Amount column + DR/CR column" ledger into a
    synthetic Date/Debit/Credit/Description/Balance CSV that LedgerParser
    already knows how to parse. This is pure column reshaping (which raw
    value goes in which synthetic column) — the actual amount/date parsing
    still happens inside LedgerParser.parse(), not here.
    """
    date_i = idx["date"]
    amount_i = idx["amount"]
    drcr_i = idx["drcr"]
    desc_i = idx.get("description") if description_key else None
    bal_i = idx.get("balance") if balance_key else None

    def cell(row: list, i: int | None):
        if i is None or i >= len(row):
            return None
        return row[i]

    out_rows: list[list[str]] = [["Date", "Debit", "Credit", "Description", "Balance"]]

    for row in rows[1:]:
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue

        date_val = cell(row, date_i)
        if isinstance(date_val, (datetime.date, datetime.datetime)):
            date_str = date_val.strftime("%d/%m/%Y")
        else:
            date_str = "" if date_val is None else str(date_val)

        amount_val = cell(row, amount_i)
        amount_str = "" if amount_val is None else str(amount_val)
        flag_val = str(cell(row, drcr_i) or "").strip().lower()

        debit_str, credit_str = "", ""
        if flag_val.startswith("d"):
            debit_str = amount_str
        elif flag_val.startswith("c"):
            credit_str = amount_str
        else:
            # No recognizable DR/CR flag on this row: best-effort fall back
            # to the sign of the raw amount text (negative -> debit).
            stripped = amount_str.strip()
            if stripped.startswith("-") or (stripped.startswith("(") and stripped.endswith(")")):
                debit_str = amount_str
            else:
                credit_str = amount_str

        desc_val = cell(row, desc_i) if desc_i is not None else None
        bal_val = cell(row, bal_i) if bal_i is not None else None

        out_rows.append(
            [
                date_str,
                debit_str,
                credit_str,
                "" if desc_val is None else str(desc_val),
                "" if bal_val is None else str(bal_val),
            ]
        )

    buf = io.StringIO()
    csv.writer(buf).writerows(out_rows)
    csv_bytes = buf.getvalue().encode("utf-8")

    col_mapping = {"date": "Date", "debit": "Debit", "credit": "Credit"}
    if desc_i is not None:
        col_mapping["description"] = "Description"
    if bal_i is not None:
        col_mapping["balance"] = "Balance"

    return csv_bytes, col_mapping


def _masked_preview_df(header_labels: list[str], preview_rows: list[list]) -> pd.DataFrame:
    safe_rows = []
    for row in preview_rows:
        row = list(row) if row is not None else []
        # Pad short rows so every preview row lines up with the header.
        row = row + [None] * (len(header_labels) - len(row))
        safe_rows.append(
            [mask_pii("" if c is None else str(c)) for c in row[: len(header_labels)]]
        )
    return pd.DataFrame(safe_rows, columns=header_labels)


# ---------------------------------------------------------------------------
# Ledger upload + column-mapping flow (CLAUDE.md UI Requirements,
# ui-agent.md rule 4)
# ---------------------------------------------------------------------------


def _handle_ledger_upload(ledger_file) -> list[StandardTransaction] | None:
    file_bytes = ledger_file.getvalue()
    file_id = f"{ledger_file.name}|{hashlib.md5(file_bytes).hexdigest()}"

    if st.session_state.get("ledger_file_id") != file_id:
        # A genuinely new/different ledger file: forget any previously
        # confirmed mapping and force the mapping UI to reappear.
        st.session_state["ledger_file_id"] = file_id
        st.session_state.pop("ledger_mapping", None)
        st.session_state["show_ledger_mapping_form"] = True

    try:
        rows = _read_raw_rows(file_bytes, ledger_file.name)
    except Exception as exc:
        st.error(f"Could not read the ledger file: {exc}")
        return None

    if not rows:
        st.error("This ledger file appears to be empty.")
        return None

    header_row = rows[0]
    headers = _header_labels(header_row)
    preview_rows = rows[1:6]

    show_form = st.session_state.get("show_ledger_mapping_form", True)

    if show_form:
        st.subheader("Map Your Ledger Columns")
        st.caption(
            "Tell us which column is which in your ledger file. This is "
            "remembered for this session so you won't be asked again for "
            "the same file."
        )
        st.dataframe(_masked_preview_df(headers, preview_rows), use_container_width=True)

        # Widget keys are scoped to this specific file (by its id) so that a
        # previously-selected value from a *different* ledger file's header
        # list can never leak in and crash a selectbox whose options just
        # changed underneath it.
        key_suffix = file_id.split("|", 1)[-1][:12]

        mode = st.radio(
            "Ledger amount format", [MODE_SEPARATE, MODE_AMOUNT_DRCR], key=f"ledger_mode_radio_{key_suffix}"
        )

        date_col = st.selectbox("Date column", headers, key=f"map_date_{key_suffix}")
        options = [NONE_OPTION] + headers

        debit_col = credit_col = amount_col = drcr_col = None
        if mode == MODE_SEPARATE:
            debit_col = st.selectbox("Debit column", options, key=f"map_debit_{key_suffix}")
            credit_col = st.selectbox("Credit column", options, key=f"map_credit_{key_suffix}")
        else:
            amount_col = st.selectbox("Amount column", headers, key=f"map_amount_{key_suffix}")
            drcr_col = st.selectbox("DR/CR column", headers, key=f"map_drcr_{key_suffix}")

        description_col = st.selectbox("Description column", options, key=f"map_description_{key_suffix}")
        balance_col = st.selectbox("Balance column", options, key=f"map_balance_{key_suffix}")

        can_confirm = True
        if mode == MODE_SEPARATE and debit_col == NONE_OPTION and credit_col == NONE_OPTION:
            st.warning("Please map at least one of Debit or Credit.")
            can_confirm = False

        if st.button("Confirm column mapping", disabled=not can_confirm):
            st.session_state["ledger_mapping"] = {
                "mode": mode,
                "date": date_col,
                "debit": None if debit_col in (None, NONE_OPTION) else debit_col,
                "credit": None if credit_col in (None, NONE_OPTION) else credit_col,
                "amount": amount_col,
                "drcr": drcr_col,
                "description": None if description_col == NONE_OPTION else description_col,
                "balance": None if balance_col == NONE_OPTION else balance_col,
            }
            st.session_state["show_ledger_mapping_form"] = False
            st.rerun()
        return None

    mapping = st.session_state.get("ledger_mapping")
    if mapping is None:
        # Defensive fallback: form was hidden but nothing saved yet.
        st.session_state["show_ledger_mapping_form"] = True
        st.rerun()
        return None

    st.success(f"Using saved ledger column mapping ({mapping['mode']}).")
    if st.button("Change column mapping"):
        st.session_state["show_ledger_mapping_form"] = True
        st.rerun()

    try:
        idx = {name: i for i, name in enumerate(headers)}
        if mapping["mode"] == MODE_SEPARATE:
            col_mapping = {"date": mapping["date"]}
            if mapping.get("debit"):
                col_mapping["debit"] = mapping["debit"]
            if mapping.get("credit"):
                col_mapping["credit"] = mapping["credit"]
            if mapping.get("description"):
                col_mapping["description"] = mapping["description"]
            if mapping.get("balance"):
                col_mapping["balance"] = mapping["balance"]
            return LedgerParser().parse(file_bytes, ledger_file.name, col_mapping)

        map_idx = {
            "date": idx[mapping["date"]],
            "amount": idx[mapping["amount"]],
            "drcr": idx[mapping["drcr"]],
        }
        if mapping.get("description"):
            map_idx["description"] = idx[mapping["description"]]
        if mapping.get("balance"):
            map_idx["balance"] = idx[mapping["balance"]]

        synthetic_bytes, col_mapping = _build_amount_drcr_csv(
            rows, map_idx, mapping.get("description"), mapping.get("balance")
        )
        return LedgerParser().parse(synthetic_bytes, "ledger_synthetic.csv", col_mapping)
    except Exception as exc:
        st.error(
            f"Could not parse the ledger with the current column mapping: {exc}. "
            "Try 'Change column mapping' above."
        )
        return None


# ---------------------------------------------------------------------------
# Results rendering
# ---------------------------------------------------------------------------


def _record_to_row(record: MatchRecord) -> dict:
    bank, ledger = record.bank_txn, record.ledger_txn
    return {
        "Tier": record.tier,
        "Bank Date": bank.date.isoformat() if bank else "",
        "Bank Description": mask_pii(bank.description) if bank else "",
        "Bank Debit": float(bank.debit) if bank else None,
        "Bank Credit": float(bank.credit) if bank else None,
        "Ledger Date": ledger.date.isoformat() if ledger else "",
        "Ledger Description": mask_pii(ledger.description) if ledger else "",
        "Ledger Debit": float(ledger.debit) if ledger else None,
        "Ledger Credit": float(ledger.credit) if ledger else None,
        "Amount Diff": float(record.amount_diff),
        "Date Diff (days)": record.date_diff_days,
        "Desc Score": record.description_score,
        "Reason": record.reason,
        "Bank Row #": bank.source_row if bank else None,
        "Ledger Row #": ledger.source_row if ledger else None,
    }


def _records_to_df(records: list[MatchRecord]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame(columns=TABLE_COLUMNS)
    return pd.DataFrame([_record_to_row(r) for r in records], columns=TABLE_COLUMNS)


def _style_by_tier(df: pd.DataFrame):
    def _row_style(row):
        color = TIER_COLORS.get(row.get("Tier"), "")
        style = f"background-color: {color}" if color else ""
        return [style] * len(row)

    return df.style.apply(_row_style, axis=1)


def _show_table(records: list[MatchRecord]) -> None:
    df = _records_to_df(records)
    if df.empty:
        st.info("No records in this category.")
        return
    st.dataframe(_style_by_tier(df), use_container_width=True)


def _matched_amount(result: ReconciliationResult) -> Decimal:
    total = Decimal("0")
    for record in result.exact + result.tolerance + result.review:
        if record.bank_txn is not None:
            total += abs(record.bank_txn.credit - record.bank_txn.debit)
    return total


def _render_summary(result: ReconciliationResult, bank_txns: list[StandardTransaction]) -> None:
    st.subheader("Summary")
    total_bank = len(bank_txns)
    matched_pct = (result.matched_count / total_bank * 100) if total_bank else 0.0
    matched_amount = _matched_amount(result)

    cols = st.columns(6)
    cols[0].metric("Matched %", f"{matched_pct:.1f}%")
    cols[1].metric("Exact", len(result.exact))
    cols[2].metric("Tolerance", len(result.tolerance))
    cols[3].metric("Review", len(result.review))
    cols[4].metric("Unmatched", result.unmatched_count)
    cols[5].metric("Total Matched Amount", f"Rs {matched_amount:,.2f}")


def _render_tabs(result: ReconciliationResult) -> None:
    st.subheader("Detailed Results")
    tab_all, tab_exact, tab_tol, tab_review, tab_bank_only, tab_ledger_only = st.tabs(
        ["All", "Exact", "Tolerance", "Review", "Bank-only", "Ledger-only"]
    )
    with tab_all:
        _show_table(result.all_records)
    with tab_exact:
        _show_table(result.exact)
    with tab_tol:
        _show_table(result.tolerance)
    with tab_review:
        _show_table(result.review)
    with tab_bank_only:
        _show_table(result.bank_only)
    with tab_ledger_only:
        _show_table(result.ledger_only)


def _render_export_section() -> None:
    st.subheader("Export")
    if st.button("Export to Excel"):
        st.info("Excel export coming in M4")


# ---------------------------------------------------------------------------
# Sidebar (uploads + settings)
# ---------------------------------------------------------------------------


def _sidebar() -> tuple:
    st.sidebar.header("1. Upload Files")
    bank_names = list(BANK_ADAPTERS.keys())
    bank_name = st.sidebar.selectbox("Bank", bank_names)
    bank_file = st.sidebar.file_uploader(
        "Bank statement (PDF, XLSX, or CSV)", type=["pdf", "xlsx", "xls", "csv"], key="bank_file"
    )
    ledger_file = st.sidebar.file_uploader(
        "Ledger (XLSX or CSV)", type=["xlsx", "xls", "csv"], key="ledger_file"
    )

    st.sidebar.header("2. Matching Settings")
    amount_tol_abs = st.sidebar.number_input(
        "Amount tolerance — Rs (absolute)", min_value=0.0, value=2.0, step=0.5
    )
    amount_tol_pct = st.sidebar.slider(
        "Amount tolerance — % (relative)", min_value=0.0, max_value=5.0, value=0.5, step=0.1
    )
    date_window = st.sidebar.slider("Date match window (± days)", min_value=0, max_value=15, value=3)
    review_date_window = st.sidebar.slider(
        "Review date window (± days)",
        min_value=date_window,
        max_value=30,
        value=max(10, date_window),
    )
    desc_threshold = st.sidebar.slider(
        "Description match threshold (Review)", min_value=50, max_value=100, value=85
    )

    st.sidebar.header("3. AI Assist")
    ai_enabled = st.sidebar.toggle(
        "Enable AI description matching",
        value=False,
        help="Off by default. Only refines already-shortlisted candidates; "
        "wired up in a later milestone.",
    )
    st.session_state["ai_description_matching_enabled"] = ai_enabled

    try:
        config = MatchConfig(
            amount_tolerance_abs=Decimal(str(amount_tol_abs)),
            amount_tolerance_pct=Decimal(str(amount_tol_pct)) / Decimal("100"),
            date_window_days=int(date_window),
            review_date_window_days=int(review_date_window),
            description_review_threshold=int(desc_threshold),
        )
    except InvalidOperation:
        st.sidebar.error("Invalid tolerance value; using defaults.")
        config = MatchConfig()

    return bank_name, bank_file, ledger_file, config


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    st.set_page_config(page_title="Bank Reconciliation", layout="wide")
    st.title("Bank Reconciliation")
    st.caption("Upload a bank statement and your ledger to generate a reconciliation report.")

    bank_name, bank_file, ledger_file, config = _sidebar()

    bank_txns: list[StandardTransaction] | None = None
    if bank_file is not None:
        try:
            parser = get_parser(bank_name)
            bank_txns = parser.parse(bank_file.getvalue(), bank_file.name)
            if not bank_txns:
                st.warning(
                    "No transactions could be extracted from the bank statement. "
                    "Double-check the file and bank selection."
                )
        except ScannedPdfError:
            st.error(SCANNED_PDF_MESSAGE)
        except Exception as exc:
            st.error(f"Could not parse the bank statement: {exc}")

    ledger_txns: list[StandardTransaction] | None = None
    if ledger_file is not None:
        ledger_txns = _handle_ledger_upload(ledger_file)

    if bank_txns and ledger_txns:
        result = reconcile(bank_txns, ledger_txns, config)
        _render_summary(result, bank_txns)
        _render_tabs(result)
        _render_export_section()
    else:
        st.info("Upload both a bank statement and a ledger file to see the reconciliation.")


main()
