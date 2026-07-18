"""Generic fuzzy-column-detection parser — the fallback used when the user
picks "Other/Auto-detect" for a bank statement (F1.3).

Handles CSV and XLSX/XLS input, per skills/bank-statement-parser/SKILL.md:
- Header-Row Detection: scans the first 25 rows for the first one where
  >= 3 cells fuzzy-match the column synonym table.
- Column Fuzzy-Matching Table (rapidfuzz ratio >= 80).
- Single `amount` + `drcr flag` column normalization into debit/credit.
- Drops junk rows (unparseable date AND both amounts zero).
- Sorts the result by date, then source_row.
"""

from __future__ import annotations

import csv
import io
from decimal import Decimal
from typing import Any

import openpyxl
from rapidfuzz import fuzz

from app.parsers.base import BaseParser, StandardTransaction, parse_amount, parse_date

# Column Fuzzy-Matching Table (skills/bank-statement-parser/SKILL.md)
SYNONYMS: dict[str, list[str]] = {
    "date": ["date", "txn date", "value date", "transaction date", "posting date", "tarikh"],
    "debit": ["debit", "withdrawal", "dr", "debit amount", "paid out", "withdrawals"],
    "credit": ["credit", "deposit", "cr", "credit amount", "paid in", "deposits"],
    "description": [
        "description", "particulars", "narration", "details",
        "transaction details", "remarks",
    ],
    "balance": ["balance", "running balance", "closing balance"],
    "amount": ["amount", "txn amount", "transaction amount"],
    "drcr": ["dr/cr", "type", "txn type", "indicator"],
}

FUZZY_THRESHOLD = 80
MAX_HEADER_SCAN_ROWS = 25


def _read_rows(file_bytes: bytes, filename: str) -> list[list[Any]]:
    """Read a CSV or XLSX/XLS file into a list of raw rows (native types
    preserved where possible, so parse_date/parse_amount can handle Excel
    serial numbers and native datetimes directly).
    """
    lower = filename.lower()
    if lower.endswith(".csv"):
        text = file_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        return [list(row) for row in reader]

    if lower.endswith(".xlsx") or lower.endswith(".xlsm") or lower.endswith(".xls"):
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        sheet = wb[wb.sheetnames[0]]
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        return rows

    raise ValueError(f"Unsupported file type for generic parser: {filename}")


def _match_target(cell_text: str) -> tuple[str | None, float]:
    """Best-matching synonym-table target for a single header cell, or
    (None, score) if nothing scores >= FUZZY_THRESHOLD.
    """
    cell_norm = cell_text.strip().lower()
    if not cell_norm:
        return None, 0.0

    best_target: str | None = None
    best_score = 0.0
    for target, synonyms in SYNONYMS.items():
        for syn in synonyms:
            score = fuzz.ratio(cell_norm, syn)
            if score > best_score:
                best_score = score
                best_target = target

    if best_score >= FUZZY_THRESHOLD:
        return best_target, best_score
    return None, best_score


def detect_header(rows: list[list[Any]]) -> tuple[int, dict[str, int], list[str]]:
    """Scan the first 25 rows for the header row (first row with >= 3
    fuzzy-matching columns). Returns (header_row_index, col_map, headers).
    """
    max_scan = min(MAX_HEADER_SCAN_ROWS, len(rows))
    for row_idx in range(max_scan):
        row = rows[row_idx]
        best_per_target: dict[str, tuple[int, float]] = {}
        match_count = 0
        for col_idx, cell in enumerate(row):
            cell_text = "" if cell is None else str(cell)
            target, score = _match_target(cell_text)
            if target:
                match_count += 1
                if target not in best_per_target or score > best_per_target[target][1]:
                    best_per_target[target] = (col_idx, score)
        if match_count >= 3:
            col_map = {target: idx for target, (idx, _score) in best_per_target.items()}
            headers = ["" if c is None else str(c) for c in row]
            return row_idx, col_map, headers

    raise ValueError(
        "Could not detect a bank-statement header row (>= 3 recognizable "
        "columns) in the first 25 rows. This file may not be a supported "
        "bank statement format."
    )


def _cell(row: list[Any], idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _extract_debit_credit(row: list[Any], col_map: dict[str, int]) -> tuple[Decimal, Decimal]:
    """Normalize whatever debit/credit representation this file uses into
    a (debit, credit) pair of non-negative Decimals.
    """
    has_debit = "debit" in col_map
    has_credit = "credit" in col_map
    has_amount = "amount" in col_map
    has_drcr = "drcr" in col_map

    if has_debit or has_credit:
        debit = parse_amount(_cell(row, col_map.get("debit")))
        credit = parse_amount(_cell(row, col_map.get("credit")))
        return abs(debit), abs(credit)

    if has_amount:
        amount = parse_amount(_cell(row, col_map["amount"]))
        if has_drcr:
            flag = str(_cell(row, col_map["drcr"]) or "").strip().lower()
            if flag.startswith("d"):
                return abs(amount), Decimal("0")
            if flag.startswith("c"):
                return Decimal("0"), abs(amount)
        # No usable flag (or flag didn't say DR/CR) — fall back to sign.
        if amount < 0:
            return abs(amount), Decimal("0")
        return Decimal("0"), amount

    return Decimal("0"), Decimal("0")


class GenericParser(BaseParser):
    """Fallback parser for "Other/Auto-detect" — fuzzy column detection on
    CSV/XLSX files with unknown bank layouts.
    """

    bank_name = "Other/Auto-detect"

    def can_parse(self, file_bytes: bytes, filename: str) -> bool:
        lower = filename.lower()
        if not (lower.endswith(".csv") or lower.endswith(".xlsx") or lower.endswith(".xls")):
            return False
        try:
            rows = _read_rows(file_bytes, filename)
            detect_header(rows)
        except Exception:
            return False
        return True

    def parse(self, file_bytes: bytes, filename: str) -> list[StandardTransaction]:
        rows = _read_rows(file_bytes, filename)
        header_idx, col_map, headers = detect_header(rows)

        def header_name(idx: int) -> str:
            if idx < len(headers) and headers[idx]:
                return headers[idx]
            return f"col_{idx}"

        transactions: list[StandardTransaction] = []

        for row_idx in range(header_idx + 1, len(rows)):
            row = rows[row_idx]
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue  # fully blank row

            source_row = row_idx + 1  # 1-based row number in the original file

            date_raw = _cell(row, col_map.get("date"))
            try:
                date_val = parse_date(date_raw)
            except Exception:
                date_val = None

            debit, credit = _extract_debit_credit(row, col_map)

            # Validation Before Returning (SKILL.md): drop junk rows where
            # the date failed to parse AND both amounts are zero.
            if date_val is None and debit == 0 and credit == 0:
                continue

            if date_val is None:
                # No valid date to build a StandardTransaction from; skip
                # rather than fabricate a date. (Amounts non-zero here means
                # a malformed row we cannot safely place on the timeline.)
                continue

            description_raw = _cell(row, col_map.get("description"))
            description = "" if description_raw is None else str(description_raw).strip()

            balance_val: Decimal | None = None
            if "balance" in col_map:
                balance_val = parse_amount(_cell(row, col_map["balance"]))

            raw_dict = {header_name(i): row[i] for i in range(len(row))}

            transactions.append(
                StandardTransaction(
                    date=date_val,
                    description=description,
                    debit=debit,
                    credit=credit,
                    balance=balance_val,
                    source_row=source_row,
                    raw=raw_dict,
                )
            )

        transactions.sort(key=lambda t: (t.date, t.source_row))
        return transactions
