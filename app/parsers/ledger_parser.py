"""Parser for the user's local ledger (F1.5).

Unlike bank_generic.py (which auto-detects columns via fuzzy matching), the
ledger is parsed using an explicit column mapping the user supplies once via
the UI's column-mapping step, e.g.:

    {"date": "Txn Date", "debit": "Withdrawal", "credit": "Deposit",
     "description": "Details", "balance": "Closing Balance"}

The first row of the file is treated as the header row (no junk-row
scanning — ledgers are the user's own, already-clean export). Mapping keys
are matched against header cells case-insensitively (exact match after
stripping whitespace).
"""

from __future__ import annotations

import csv
import io
from decimal import Decimal
from typing import Any

import openpyxl

from app.parsers.base import StandardTransaction, parse_amount, parse_date

REQUIRED_MAPPING_KEYS = ("date",)
OPTIONAL_MAPPING_KEYS = ("debit", "credit", "description", "balance")


def _read_rows(file_bytes: bytes, filename: str) -> list[list[Any]]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        text = file_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        return [list(row) for row in reader]

    if lower.endswith(".xlsx") or lower.endswith(".xlsm") or lower.endswith(".xls"):
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        sheet = wb[wb.sheetnames[0]]
        return [list(row) for row in sheet.iter_rows(values_only=True)]

    raise ValueError(f"Unsupported file type for ledger parser: {filename}")


def _cell(row: list[Any], idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


class LedgerParser:
    """Parses the local ledger file into StandardTransaction using a
    user-supplied column mapping rather than auto-detection.
    """

    def parse(
        self,
        file_bytes: bytes,
        filename: str,
        column_mapping: dict[str, str],
    ) -> list[StandardTransaction]:
        if "date" not in column_mapping:
            raise ValueError("column_mapping must include a 'date' entry")
        if "debit" not in column_mapping and "credit" not in column_mapping:
            raise ValueError(
                "column_mapping must include at least one of 'debit' or 'credit'"
            )

        rows = _read_rows(file_bytes, filename)
        if not rows:
            return []

        header_row = ["" if c is None else str(c).strip() for c in rows[0]]
        header_lookup = {h.lower(): idx for idx, h in enumerate(header_row) if h}

        col_map: dict[str, int] = {}
        for key, header_name in column_mapping.items():
            if key not in REQUIRED_MAPPING_KEYS + OPTIONAL_MAPPING_KEYS:
                continue
            lookup_key = header_name.strip().lower()
            if lookup_key not in header_lookup:
                raise ValueError(
                    f"Mapped column {header_name!r} for {key!r} not found in "
                    f"ledger header row: {header_row!r}"
                )
            col_map[key] = header_lookup[lookup_key]

        transactions: list[StandardTransaction] = []

        for row_idx in range(1, len(rows)):
            row = rows[row_idx]
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue  # fully blank row

            source_row = row_idx + 1  # 1-based row number in the original file

            date_raw = _cell(row, col_map.get("date"))
            try:
                date_val = parse_date(date_raw)
            except Exception:
                date_val = None

            debit = parse_amount(_cell(row, col_map.get("debit"))) if "debit" in col_map else Decimal("0")
            credit = parse_amount(_cell(row, col_map.get("credit"))) if "credit" in col_map else Decimal("0")
            debit, credit = abs(debit), abs(credit)

            # Drop junk rows: unparseable date AND both amounts zero.
            if date_val is None and debit == 0 and credit == 0:
                continue

            if date_val is None:
                # Cannot build a valid StandardTransaction without a date.
                continue

            description_raw = _cell(row, col_map.get("description"))
            description = "" if description_raw is None else str(description_raw).strip()

            balance_val: Decimal | None = None
            if "balance" in col_map:
                balance_val = parse_amount(_cell(row, col_map["balance"]))

            raw_dict = {header_row[i] or f"col_{i}": row[i] for i in range(len(row))}

            transactions.append(
                StandardTransaction(
                    date=date_val,
                    description=description,
                    debit=debit,
                    credit=credit,
                    balance=balance_val,
                    source_row=source_row,
                    raw=raw_dict,
                )
            )

        transactions.sort(key=lambda t: (t.date, t.source_row))
        return transactions
