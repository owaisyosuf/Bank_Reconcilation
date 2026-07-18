"""Tests for app/export/excel_report.py — the color-coded Excel export (M4).

Builds a small synthetic `ReconciliationResult` covering every tier, then
loads the produced bytes back with openpyxl to assert on sheet names,
summary counts, fill colors, and PII masking.
"""

import datetime
import io
from decimal import Decimal

import openpyxl

from app.export.excel_report import TIER_COLORS, build_excel_report
from app.matching.engine import ReconciliationResult
from app.matching.scoring import MatchRecord
from app.parsers.base import StandardTransaction

BASE_DATE = datetime.date(2026, 1, 15)


def make_txn(description: str, amount, source_row: int, date=BASE_DATE) -> StandardTransaction:
    amount = Decimal(str(amount))
    if amount >= 0:
        credit, debit = amount, Decimal("0")
    else:
        credit, debit = Decimal("0"), -amount
    return StandardTransaction(
        date=date,
        description=description,
        debit=debit,
        credit=credit,
        balance=None,
        source_row=source_row,
        raw={},
    )


def build_sample_result() -> ReconciliationResult:
    result = ReconciliationResult()

    # EXACT: one record, with a CNIC-like description to verify masking.
    result.exact.append(
        MatchRecord(
            tier="EXACT",
            bank_txn=make_txn("Transfer from 42101-1234567-1", 15000, 1),
            ledger_txn=make_txn("Transfer from 42101-1234567-1", 15000, 1),
            amount_diff=Decimal("0"),
            date_diff_days=0,
            description_score=100,
            reason="exact",
        )
    )

    # TOLERANCE: Rs 1.50 difference.
    result.tolerance.append(
        MatchRecord(
            tier="TOLERANCE",
            bank_txn=make_txn("Bank charges", 999, 2),
            ledger_txn=make_txn("Bank charges", 1000.50, 2),
            amount_diff=Decimal("-1.50"),
            date_diff_days=0,
            description_score=95,
            reason="within_tolerance",
        )
    )

    # REVIEW: date offset.
    result.review.append(
        MatchRecord(
            tier="REVIEW",
            bank_txn=make_txn("Cheque clearing", 5000, 3, date=BASE_DATE),
            ledger_txn=make_txn("Cheque clearing", 5000, 3, date=BASE_DATE + datetime.timedelta(days=5)),
            amount_diff=Decimal("0"),
            date_diff_days=-5,
            description_score=100,
            reason="date_offset_5d",
        )
    )

    # Bank only (unmatched).
    result.bank_only.append(
        MatchRecord(
            tier="UNMATCHED",
            bank_txn=make_txn("POS purchase", -750, 4),
            ledger_txn=None,
            amount_diff=Decimal("-750"),
            date_diff_days=0,
            description_score=0,
            reason="no_match_bank_only",
        )
    )

    # Ledger only (unmatched).
    result.ledger_only.append(
        MatchRecord(
            tier="UNMATCHED",
            bank_txn=None,
            ledger_txn=make_txn("Manual journal entry", 2500, 5),
            amount_diff=Decimal("2500"),
            date_diff_days=0,
            description_score=0,
            reason="no_match_ledger_only",
        )
    )

    return result


def test_all_sheets_present():
    result = build_sample_result()
    wb = openpyxl.load_workbook(io.BytesIO(build_excel_report(result)), data_only=True)

    expected = ["Summary", "Exact", "Tolerance", "Review", "Bank Only", "Ledger Only"]
    assert wb.sheetnames == expected


def test_summary_sheet_counts():
    result = build_sample_result()
    wb = openpyxl.load_workbook(io.BytesIO(build_excel_report(result)), data_only=True)
    ws = wb["Summary"]

    rows = {row[0]: row for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}

    assert rows["Exact"][1] == 1
    assert rows["Tolerance"][1] == 1
    assert rows["Review"][1] == 1
    assert rows["Bank Only (unmatched)"][1] == 1
    assert rows["Ledger Only (unmatched)"][1] == 1

    assert rows["Total Matched Count"][1] == result.matched_count
    assert rows["Total Unmatched Count"][1] == result.unmatched_count
    assert "Generated At" in rows


def test_tier_sheet_fill_color():
    result = build_sample_result()
    wb = openpyxl.load_workbook(io.BytesIO(build_excel_report(result)))

    ws = wb["Exact"]
    data_cell = ws.cell(row=2, column=1)
    assert data_cell.fill.start_color.rgb.endswith(TIER_COLORS["EXACT"])

    ws_tol = wb["Tolerance"]
    assert ws_tol.cell(row=2, column=1).fill.start_color.rgb.endswith(TIER_COLORS["TOLERANCE"])

    ws_review = wb["Review"]
    assert ws_review.cell(row=2, column=1).fill.start_color.rgb.endswith(TIER_COLORS["REVIEW"])

    ws_bank_only = wb["Bank Only"]
    assert ws_bank_only.cell(row=2, column=1).fill.start_color.rgb.endswith(TIER_COLORS["UNMATCHED"])

    ws_ledger_only = wb["Ledger Only"]
    assert ws_ledger_only.cell(row=2, column=1).fill.start_color.rgb.endswith(TIER_COLORS["UNMATCHED"])


def test_frozen_header_row_on_every_sheet():
    result = build_sample_result()
    wb = openpyxl.load_workbook(io.BytesIO(build_excel_report(result)))
    for name in wb.sheetnames:
        assert wb[name].freeze_panes == "A2"


def test_pii_masked_in_export():
    result = build_sample_result()
    wb = openpyxl.load_workbook(io.BytesIO(build_excel_report(result)), data_only=True)
    ws = wb["Exact"]

    # Bank Description is column 2.
    bank_description = ws.cell(row=2, column=2).value
    assert "42101-1234567-1" not in bank_description
    assert "4210112345671" not in bank_description
    assert bank_description.endswith("5671")


def test_traceability_source_rows_present():
    result = build_sample_result()
    wb = openpyxl.load_workbook(io.BytesIO(build_excel_report(result)), data_only=True)
    ws = wb["Exact"]

    # Bank Row # and Ledger Row # are the last two columns.
    row = [c.value for c in ws[2]]
    assert row[-2] == 1  # Bank Row #
    assert row[-1] == 1  # Ledger Row #


def test_bytes_are_valid_xlsx():
    result = build_sample_result()
    data = build_excel_report(result)
    assert isinstance(data, bytes)
    assert data[:2] == b"PK"  # xlsx is a zip archive
